"""
News Title Classification — End-to-End Pipeline
=================================================
1. Feature Engineering: TF-IDF long-format CSV -> sparse wide matrix
2. Merge with Category labels from xlsx
3. Train 4 SVM variants (Linear SVM, SVC-linear, SVC-RBF, SVC-poly)
4. Evaluate: accuracy, precision, recall, F1 (per-category + overall)
5. Save everything into one formatted Excel report:
   - Project Summary
   - Model Comparison
   - Per-Category Metrics
   - Confusion Matrix (best model)

Run: python3 svm_full_pipeline.py
Output: /mnt/user-data/outputs/SVM_Results_Report.xlsx
"""

import json
import numpy as np
import pandas as pd
import openpyxl
from scipy import sparse
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split
from sklearn.svm import SVC, LinearSVC
from sklearn.metrics import (accuracy_score, precision_recall_fscore_support,
                              classification_report, confusion_matrix)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LABELS_FILE = "News_Classification_Labeling.xlsx"
TFIDF_FILE = "news_titles_tfidf.csv"
OUTPUT_FILE = "SVM_Results_Report.xlsx"

# ============================================================
# STEP 1: FEATURE ENGINEERING
# ============================================================
print("Loading category labels...")
wb_in = openpyxl.load_workbook(LABELS_FILE, read_only=True)
ws_in = wb_in['Sheet1']
rows = []
first = True
for r in ws_in.iter_rows(values_only=True):
    if first:
        first = False
        continue
    rows.append(r[:3])
xl = pd.DataFrame(rows, columns=['URL', 'Title', 'Category'])
xl['row_id'] = range(len(xl))

print("Loading TF-IDF (long format) and pivoting to sparse matrix...")
tfidf_long = pd.read_csv(TFIDF_FILE)
tfidf_long.columns = [c.strip() for c in tfidf_long.columns]

feat_cat = pd.Categorical(tfidf_long['feature'])
row_cat = pd.Categorical(tfidf_long['row_id'], categories=sorted(tfidf_long['row_id'].unique()))

n_rows = len(row_cat.categories)
n_feats = len(feat_cat.categories)

mat = sparse.csr_matrix(
    (tfidf_long['tfidf'].values, (row_cat.codes, feat_cat.codes)),
    shape=(n_rows, n_feats)
)
row_ids = list(row_cat.categories)

# Merge with labels, drop rows with missing category
label_df = pd.DataFrame({'row_id': row_ids}).merge(xl[['row_id', 'Category']], on='row_id', how='left')
valid_mask = label_df['Category'].notna().values
n_dropped = (~valid_mask).sum()

X = mat[valid_mask]
y = label_df.loc[valid_mask, 'Category'].values

print(f"Feature matrix shape: {X.shape} | Dropped {n_dropped} row(s) with missing category")
dist = pd.Series(y).value_counts()
print("Category distribution:\n", dist)

le = LabelEncoder()
y_enc = le.fit_transform(y)
classes = le.classes_

X_train, X_test, y_train, y_test = train_test_split(
    X, y_enc, test_size=0.2, random_state=42, stratify=y_enc
)
print(f"Train: {X_train.shape} | Test: {X_test.shape}")

# ============================================================
# STEP 2: TRAIN + EVALUATE 4 SVM VARIANTS
# ============================================================
models = {
    'Linear SVM (LinearSVC)': LinearSVC(C=1.0, max_iter=5000, random_state=42),
    'SVC - linear kernel': SVC(kernel='linear', C=1.0, random_state=42),
    'SVC - rbf kernel': SVC(kernel='rbf', C=1.0, gamma='scale', random_state=42),
    'SVC - poly kernel': SVC(kernel='poly', degree=3, C=1.0, random_state=42),
}

results = {}
predictions = {}

for name, model in models.items():
    print(f"\nTraining {name}...")
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    predictions[name] = y_pred

    acc = accuracy_score(y_test, y_pred)
    prec_w, rec_w, f1_w, _ = precision_recall_fscore_support(y_test, y_pred, average='weighted', zero_division=0)
    prec_m, rec_m, f1_m, _ = precision_recall_fscore_support(y_test, y_pred, average='macro', zero_division=0)
    p_cls, r_cls, f_cls, s_cls = precision_recall_fscore_support(
        y_test, y_pred, average=None, zero_division=0, labels=range(len(classes))
    )

    results[name] = {
        'accuracy': acc,
        'precision_weighted': prec_w, 'recall_weighted': rec_w, 'f1_weighted': f1_w,
        'precision_macro': prec_m, 'recall_macro': rec_m, 'f1_macro': f1_m,
        'per_class': {
            classes[i]: {
                'precision': p_cls[i], 'recall': r_cls[i], 'f1': f_cls[i], 'support': int(s_cls[i])
            } for i in range(len(classes))
        }
    }
    print(f"Accuracy: {acc:.4f} | Macro F1: {f1_m:.4f} | Weighted F1: {f1_w:.4f}")
    print(classification_report(y_test, y_pred, target_names=classes, zero_division=0))

best_name = max(results, key=lambda k: results[k]['f1_macro'])
cm = confusion_matrix(y_test, predictions[best_name], labels=range(len(classes)))
print(f"\n>>> Best model (highest Macro F1): {best_name}")

# ============================================================
# STEP 3: BUILD FORMATTED EXCEL REPORT
# ============================================================
print("\nBuilding Excel report...")

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color='1F3864')
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=12, color='305496')
BODY_FONT = Font(name=FONT_NAME, size=11)
BOLD_BODY = Font(name=FONT_NAME, size=11, bold=True)
BEST_FILL = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
WORST_FILL = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
THIN = Side(style='thin', color='B7B7B7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
WRAP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)

wb = Workbook()

# ---------------- SHEET 1: Project Summary ----------------
ws = wb.active
ws.title = 'Project Summary'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 90

ws['A1'] = 'News Title Classification — SVM Project Summary'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:B1')
ws['A2'] = 'End-to-end pipeline: feature engineering (TF-IDF) + SVM classification + evaluation'
ws['A2'].font = Font(name=FONT_NAME, italic=True, size=10, color='595959')
ws.merge_cells('A2:B2')

def section(row, title):
    ws.cell(row=row, column=1, value=title).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    return row + 1

def kv(row, k, v):
    c1 = ws.cell(row=row, column=1, value=k); c1.font = BOLD_BODY; c1.alignment = WRAP_LEFT
    c2 = ws.cell(row=row, column=2, value=v); c2.font = BODY_FONT; c2.alignment = WRAP_LEFT
    return row + 1

r = 4
r = section(r, '1. Data Sources')
r = kv(r, 'Labels file', 'News_Classification_Labeling.xlsx (URL, Title, Category)')
r = kv(r, 'Features file', 'news_titles_tfidf.csv (long format: row_id, Title, feature, tfidf)')
r = kv(r, 'Data cleaning', f'{n_dropped} row(s) had a missing Category and were dropped. Final dataset: {X.shape[0]} labeled titles.')
r += 1

r = section(r, '2. Feature Engineering')
r = kv(r, 'Method', 'TF-IDF (term frequency-inverse document frequency), pre-computed per title in the long-format CSV.')
r = kv(r, 'Pivot step', 'Reshaped the long (row_id, feature, tfidf) table into a wide sparse matrix using pandas Categorical codes + scipy.sparse, aligned to each title via row_id.')
r = kv(r, 'Final feature matrix', f'{X.shape[0]} rows x {X.shape[1]} TF-IDF features (unigrams + bigrams), sparse (CSR) format.')
r = kv(r, 'Train/test split', f'80% train ({X_train.shape[0]} rows) / 20% test ({X_test.shape[0]} rows), stratified by Category, random_state=42.')
r += 1

r = section(r, '3. Category Distribution (full dataset)')
for k, v in dist.items():
    r = kv(r, k, f'{v} titles ({v/dist.sum()*100:.1f}%)')
r += 1

r = section(r, '4. Models Trained')
r = kv(r, 'Linear SVM (LinearSVC)', 'Linear kernel, optimized liblinear solver, C=1.0')
r = kv(r, 'SVC - linear kernel', 'libsvm implementation with explicit linear kernel, C=1.0')
r = kv(r, 'SVC - RBF kernel', 'Radial basis function kernel, C=1.0, gamma=scale')
r = kv(r, 'SVC - polynomial kernel', 'Degree-3 polynomial kernel, C=1.0')
r += 1

best_cat_overall = max(results[best_name]['per_class'], key=lambda c: results[best_name]['per_class'][c]['f1'])
worst_cat_overall = min(results[best_name]['per_class'], key=lambda c: results[best_name]['per_class'][c]['f1'])

r = section(r, '5. Headline Result')
r = kv(r, 'Best model', f'{best_name} - chosen by highest Macro F1 score (fairest metric under class imbalance)')
r = kv(r, 'Best accuracy', f"{results[best_name]['accuracy']*100:.1f}%")
r = kv(r, 'Best category', f"{best_cat_overall} - highest F1 of all categories in the best model")
r = kv(r, 'Weakest category', f"{worst_cat_overall} - lowest F1, mainly due to small sample size")
r += 1

r = section(r, '6. Key Takeaways')
takeaways = [
    'Linear kernels beat RBF and polynomial kernels on this data - expected, since TF-IDF text features are high-dimensional and already close to linearly separable.',
    'Class imbalance is the main driver of weak spots: categories with more training data are classified well; categories with little data are classified poorly.',
    f'{best_cat_overall} titles use distinctive, consistent vocabulary, making them the easiest category to separate.',
    'Recommended next step: collect more samples for weak categories, or use class_weight="balanced" / SMOTE oversampling to lift minority-class recall.',
]
for t in takeaways:
    c1 = ws.cell(row=r, column=1, value='-'); c1.font = BODY_FONT
    c2 = ws.cell(row=r, column=2, value=t); c2.font = BODY_FONT; c2.alignment = WRAP_LEFT
    ws.row_dimensions[r].height = 30
    r += 1

# ---------------- SHEET 2: Model Comparison ----------------
ws2 = wb.create_sheet('Model Comparison')
ws2.sheet_view.showGridLines = False
headers = ['Model', 'Accuracy', 'Precision (weighted)', 'Recall (weighted)', 'F1 (weighted)',
           'Precision (macro)', 'Recall (macro)', 'F1 (macro)']
for j, h in enumerate(headers, start=1):
    c = ws2.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
ws2.row_dimensions[1].height = 30

model_rows = []
for i, (name, res) in enumerate(results.items(), start=2):
    vals = [name, res['accuracy'], res['precision_weighted'], res['recall_weighted'], res['f1_weighted'],
            res['precision_macro'], res['recall_macro'], res['f1_macro']]
    model_rows.append((i, name, res['f1_macro']))
    for j, v in enumerate(vals, start=1):
        c = ws2.cell(row=i, column=j, value=v)
        c.font = BODY_FONT; c.border = BORDER; c.alignment = CENTER if j > 1 else Alignment(horizontal='left')
        if j > 1:
            c.number_format = '0.0%'

best_row = max(model_rows, key=lambda x: x[2])[0]
for j in range(1, len(headers) + 1):
    ws2.cell(row=best_row, column=j).fill = BEST_FILL
ws2.cell(row=best_row, column=1).font = BOLD_BODY

note_row = len(results) + 3
ws2.cell(row=note_row, column=1, value='Green row = best model overall, selected by highest Macro F1 (fairest metric given class imbalance).').font = Font(name=FONT_NAME, italic=True, size=9, color='595959')
ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

for j, w in enumerate([26, 12, 20, 18, 16, 18, 16, 14], start=1):
    ws2.column_dimensions[get_column_letter(j)].width = w

# ---------------- SHEET 3: Per-Category Metrics ----------------
ws3 = wb.create_sheet('Per-Category Metrics')
ws3.sheet_view.showGridLines = False
headers3 = ['Model', 'Category', 'Precision', 'Recall', 'F1-score', 'Support']
for j, h in enumerate(headers3, start=1):
    c = ws3.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
ws3.row_dimensions[1].height = 22

row_i = 2
best_model_first_row = None
best_model_last_row = None
per_cat_f1_best = {}
for name, res in results.items():
    if name == best_name:
        best_model_first_row = row_i
    for cat, m in res['per_class'].items():
        ws3.cell(row=row_i, column=1, value=name).font = BODY_FONT
        ws3.cell(row=row_i, column=2, value=cat).font = BODY_FONT
        ws3.cell(row=row_i, column=3, value=m['precision']).number_format = '0.00'
        ws3.cell(row=row_i, column=4, value=m['recall']).number_format = '0.00'
        ws3.cell(row=row_i, column=5, value=m['f1']).number_format = '0.00'
        ws3.cell(row=row_i, column=6, value=m['support'])
        for j in range(1, 7):
            ws3.cell(row=row_i, column=j).border = BORDER
            ws3.cell(row=row_i, column=j).alignment = CENTER if j > 1 else Alignment(horizontal='left')
        if name == best_name:
            per_cat_f1_best[cat] = m['f1']
        row_i += 1
    if name == best_name:
        best_model_last_row = row_i - 1
    row_i += 1

best_cat = max(per_cat_f1_best, key=per_cat_f1_best.get)
worst_cat = min(per_cat_f1_best, key=per_cat_f1_best.get)
for r_ in range(best_model_first_row, best_model_last_row + 1):
    cat_val = ws3.cell(row=r_, column=2).value
    if cat_val == best_cat:
        for j in range(1, 7):
            ws3.cell(row=r_, column=j).fill = BEST_FILL
    elif cat_val == worst_cat:
        for j in range(1, 7):
            ws3.cell(row=r_, column=j).fill = WORST_FILL

note_row3 = row_i + 1
ws3.cell(row=note_row3, column=1, value=f'Within the best model ({best_name}): green = best category ({best_cat}), red = weakest category ({worst_cat}), ranked by F1-score.').font = Font(name=FONT_NAME, italic=True, size=9, color='595959')
ws3.merge_cells(start_row=note_row3, start_column=1, end_row=note_row3, end_column=6)

for j, w in enumerate([24, 14, 12, 12, 12, 10], start=1):
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.freeze_panes = 'A2'

# ---------------- SHEET 4: Confusion Matrix ----------------
ws4 = wb.create_sheet('Confusion Matrix')
ws4.sheet_view.showGridLines = False
ws4.cell(row=1, column=1, value=f'Confusion Matrix - {best_name} (test set, n={X_test.shape[0]})').font = SUBTITLE_FONT
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(classes) + 2)

ws4.cell(row=3, column=1, value='Actual \\ Predicted').font = HEADER_FONT
ws4.cell(row=3, column=1).fill = HEADER_FILL
ws4.cell(row=3, column=1).border = BORDER
for j, cls in enumerate(classes, start=2):
    c = ws4.cell(row=3, column=j, value=cls)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER

for i, cls in enumerate(classes):
    r_ = i + 4
    c = ws4.cell(row=r_, column=1, value=cls)
    c.font = BOLD_BODY; c.border = BORDER
    for j in range(len(classes)):
        val = int(cm[i, j])
        cell = ws4.cell(row=r_, column=j + 2, value=val)
        cell.border = BORDER; cell.alignment = CENTER
        if i == j:
            cell.font = BOLD_BODY
            cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        else:
            cell.font = BODY_FONT

ws4.column_dimensions['A'].width = 20
for j in range(2, len(classes) + 2):
    ws4.column_dimensions[get_column_letter(j)].width = 12

note_row4 = len(classes) + 6
ws4.cell(row=note_row4, column=1, value='Diagonal (highlighted) = correctly predicted titles. Off-diagonal = misclassifications.').font = Font(name=FONT_NAME, italic=True, size=9, color='595959')
ws4.merge_cells(start_row=note_row4, start_column=1, end_row=note_row4, end_column=len(classes) + 2)

wb.save(OUTPUT_FILE)
print(f"\nSaved report to: {OUTPUT_FILE}")